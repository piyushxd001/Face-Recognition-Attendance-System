import os
import cv2
import shutil
import threading
import numpy as np
from datetime import datetime
from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook

import tkinter as tk
from tkinter import ttk, simpledialog, messagebox

# CONFIGURATION
DATASET_DIR = "images"
RECOGNIZER_FILE = "face_recognizer.yml"
LABELS_FILE = "labels.npy"
ATTENDANCE_FILE = "attendance.xlsx"
SHEET_NAME = "Attendance"

CAPTURE_TARGET = 10
FRAME_WIDTH = 640
FRAME_HEIGHT = 480
CONFIDENCE_THRESHOLD = 85

os.makedirs(DATASET_DIR, exist_ok=True)

# ATTENDANCE FILE INITIALIZATION

def ensure_attendance_file():
    if not os.path.exists(ATTENDANCE_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Name", "Date", "Time"])
        wb.save(ATTENDANCE_FILE)

ensure_attendance_file()


# GLOBALS
running = False
capture_mode = False
capture_name = ""
capture_count = 0
label_dict = {}
marked_today = set()
training_thread = None

status_var = None
video_label = None
student_tree = None
attendance_tree = None
search_name_var = None
search_date_var = None

# Video
video = cv2.VideoCapture(0)
video.set(cv2.CAP_PROP_FRAME_WIDTH, FRAME_WIDTH)
video.set(cv2.CAP_PROP_FRAME_HEIGHT, FRAME_HEIGHT)

# Face recognizer
face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
)

try:
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer_available = True
except Exception:
    recognizer_available = False
    recognizer = None


# HELPER UTILITIES

def get_today_date():
    return datetime.now().strftime("%Y-%m-%d")

def safe_status(msg):
    if status_var:
        status_var.set(msg)
    else:
        print("STATUS:", msg)


# LOAD & SAVE RECOGNIZER

def load_trained_recognizer():
    global recognizer, label_dict

    if not recognizer_available:
        safe_status("⚠️ opencv-contrib missing.")
        return

    if not os.path.exists(RECOGNIZER_FILE):
        safe_status("ℹ️ No trained recognizer found.")
        label_dict = {}
        return

    try:
        recognizer.read(RECOGNIZER_FILE)
        if os.path.exists(LABELS_FILE):
            label_dict = np.load(LABELS_FILE, allow_pickle=True).item()
        safe_status("✅ Recognizer loaded")
    except Exception as e:
        safe_status(f"❌ Load error: {e}")
        label_dict = {}


# ATTENDANCE FUNCTIONS


def show_fading_message(text):
    msg = tk.Label(root, text=text, fg="lightgreen", bg="#071029",
                   font=("Segoe UI", 12, "bold"))
    msg.place(relx=0.5, rely=0.02, anchor="n")
    root.after(2000, msg.destroy)

def already_marked_this_hour(name):
    try:
        wb = load_workbook(ATTENDANCE_FILE)
        ws = wb[SHEET_NAME]

        today = get_today_date()
        current_hour = datetime.now().strftime("%H")

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_name, row_date, row_time = row

            if row_name == name and row_date == today:
                saved_hour = str(row_time).split(":")[0]
                if saved_hour == current_hour:
                    return True
    except:
        pass

    return False


def load_marked_today():
    global marked_today
    marked_today = set()
    today = get_today_date()

    try:
        wb = load_workbook(ATTENDANCE_FILE)
        if SHEET_NAME not in wb.sheetnames:
            return

        for row in wb[SHEET_NAME].iter_rows(min_row=2, values_only=True):
            if row and row[1] == today:
                marked_today.add(row[0])

    except Exception as e:
        safe_status(f"⚠️ Attendance read error: {e}")


def mark_attendance_excel(name):
    now = datetime.now()

    # ---- Block duplicate within same hour ----
    if already_marked_this_hour(name):
        show_fading_message(f"Already marked this hour for {name}")
        return

    try:
        wb = load_workbook(ATTENDANCE_FILE)
        ws = wb.get_sheet_by_name(SHEET_NAME)

        ws.append([name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S")])
        wb.save(ATTENDANCE_FILE)

        marked_today.add(name)
        safe_status(f"📋 Attendance marked for {name}")

        # ---- Non-blocking soft popup ----
        show_fading_message(f"Attendance registered for {name}")

    except Exception as e:
        safe_status(f"❌ Write error: {e}")



# TRAINING MODEL

def train_recognizer_threadsafe():
    global training_thread
    if training_thread and training_thread.is_alive():
        messagebox.showinfo("Training", "Training already running.")
        return

    training_thread = threading.Thread(target=train_recognizer, daemon=True)
    training_thread.start()

def train_recognizer():
    global recognizer, label_dict

    if not recognizer_available:
        safe_status("⚠️ No LBPH available.")
        return

    faces, labels = [], []
    label_dict = {}
    label_counter = 0

    # Collect dataset
    for student in sorted(os.listdir(DATASET_DIR)):
        folder = os.path.join(DATASET_DIR, student)
        if not os.path.isdir(folder):
            continue

        images = [f for f in os.listdir(folder)
                  if f.lower().endswith(('.png', '.jpg', '.jpeg'))]

        if not images:
            continue

        label_dict[label_counter] = student

        for img_file in images:
            img_path = os.path.join(folder, img_file)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            if img is None:
                continue
            img = cv2.resize(img, (200, 200))
            faces.append(img)
            labels.append(label_counter)

        label_counter += 1

    if not faces:
        safe_status("⚠️ No data for training.")
        messagebox.showwarning("Training", "No face data found.")
        return

    try:
        safe_status("⏳ Training recognizer...")
        recognizer = cv2.face.LBPHFaceRecognizer_create()
        recognizer.train(faces, np.array(labels))

        recognizer.save(RECOGNIZER_FILE)
        np.save(LABELS_FILE, label_dict)

        safe_status("✅ Training complete.")
        load_trained_recognizer()

        root.after(100, update_student_list)

    except Exception as e:
        safe_status(f"❌ Training error: {e}")


# MAIN CAMERA LOOP
def detect_faces():
    global capture_count, capture_mode

    if not running:
        return

    ret, frame = video.read()
    if not ret:
        return video_label.after(100, detect_faces)

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.1, 5, minSize=(80, 80))

    for (x, y, w, h) in faces:
        roi = gray[y:y+h, x:x+w]
        roi = cv2.resize(roi, (200, 200))

        cv2.rectangle(frame, (x, y), (x+w, y+h), (50, 150, 255), 2)

        # CAPTURE MODE
        if capture_mode:
            folder = os.path.join(DATASET_DIR, capture_name)
            os.makedirs(folder, exist_ok=True)
            filename = datetime.now().strftime("%Y%m%d%H%M%S%f") + ".jpg"
            cv2.imwrite(os.path.join(folder, filename), roi)

            capture_count += 1
            safe_status(f"📸 Capturing {capture_name}: {capture_count}/{CAPTURE_TARGET}")

            if capture_count >= CAPTURE_TARGET:
                messagebox.showinfo("Capture", f"Captured {CAPTURE_TARGET} images")
                stop_capture()
                train_recognizer_threadsafe()
                load_trained_recognizer()

        # RECOGNITION
        elif label_dict and recognizer:
            try:
                label, conf = recognizer.predict(roi)
                name = label_dict.get(label, "Unknown")

                if conf < CONFIDENCE_THRESHOLD:
                    cv2.putText(frame, f"{name} ({int(conf)})",
                                (x, y - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                (0, 220, 120), 2)
                    mark_attendance_excel(name)

                else:
                    cv2.putText(frame, f"Unknown ({int(conf)})",
                                (x, y - 10),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                                (220, 50, 50), 2)

            except:
                cv2.putText(frame, "Err", (x, y - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                            (220, 50, 50), 2)

        else:
            cv2.putText(frame, "No trained data", (x, y - 10),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.7,
                        (200, 200, 200), 2)

    # Convert to Tkinter
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    img = ImageTk.PhotoImage(Image.fromarray(rgb))
    video_label.config(image=img)
    video_label.image = img

    video_label.after(30, detect_faces)


# STUDENT CONTROLS

def start_capture():
    global capture_mode, capture_name, capture_count

    name = simpledialog.askstring("New Student", "Enter name:")
    if not name:
        return

    name = name.strip()
    if not name:
        return

    folder = os.path.join(DATASET_DIR, name)
    if os.path.exists(folder) and os.listdir(folder):
        if not messagebox.askyesno("Overwrite?", f"Add more images for '{name}'?"):
            return

    os.makedirs(folder, exist_ok=True)
    capture_mode = True
    capture_name = name
    capture_count = 0
    safe_status(f"📸 Capturing {name}...")

def stop_capture():
    global capture_mode, capture_name, capture_count
    capture_mode = False
    capture_name = ""
    capture_count = 0
    safe_status("📸 Capture stopped.")

def start_camera():
    global running
    if not video.isOpened():
        messagebox.showerror("Camera Error", "Camera not working.")
        return

    running = True
    load_trained_recognizer()
    load_marked_today()
    detect_faces()
    safe_status("🎥 Camera started")

def stop_camera():
    global running
    running = False
    safe_status("⏹ Camera stopped.")

def close_app():
    global running
    running = False
    try:
        video.release()
    except:
        pass
    root.quit()
    root.destroy()

# STUDENT MANAGEMENT

def update_student_list():
    student_tree.delete(*student_tree.get_children())
    for name in sorted(os.listdir(DATASET_DIR)):
        if os.path.isdir(os.path.join(DATASET_DIR, name)):
            student_tree.insert("", "end", values=(name,))

def delete_student():
    sel = student_tree.selection()
    if not sel:
        return

    name = student_tree.item(sel[0])['values'][0]
    if not messagebox.askyesno("Delete?", f"Delete '{name}'?"):
        return

    try:
        shutil.rmtree(os.path.join(DATASET_DIR, name))

        if os.path.exists(LABELS_FILE):
            old = np.load(LABELS_FILE, allow_pickle=True).item()
            new = {k:v for k,v in old.items() if v != name}
            np.save(LABELS_FILE, new)

        update_student_list()
        train_recognizer_threadsafe()
        safe_status(f"🗑️ Deleted {name}")

    except Exception as e:
        messagebox.showerror("Error", str(e))


# ATTENDANCE VIEW + SEARCH

def read_attendance_rows():
    rows = []
    try:
        wb = load_workbook(ATTENDANCE_FILE, read_only=True)
        ws = wb.get_sheet_by_name(SHEET_NAME)
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(row)
    except Exception as e:
        safe_status(f"⚠️ Read error: {e}")
    return rows

def show_attendance():
    attendance_tree.delete(*attendance_tree.get_children())
    rows = read_attendance_rows()

    name_filter = search_name_var.get().strip().lower()
    date_filter = search_date_var.get().strip()

    for name, date, time in rows:
        if name_filter and name_filter not in name.lower():
            continue
        if date_filter and date_filter != str(date):
            continue
        attendance_tree.insert("", "end", values=(name, date, time))

def clear_filters():
    search_name_var.set("")
    search_date_var.set("")
    show_attendance()

# GUI INITIALIZATION
root = tk.Tk()
root.title("Face Recognition Attendance System")
root.geometry("1180x760")
root.configure(bg="#0f1720")

# Styles
style = ttk.Style()
style.theme_use("clam")
style.configure("Treeview", background="#0f1720",
                fieldbackground="#0f1720", foreground="white")
style.configure("Treeview.Heading", font=("Segoe UI", 11, "bold"))

# Status bar
status_var = tk.StringVar(value="Ready")
tk.Label(root, textvariable=status_var,
         bg="#071029", fg="#C7D2FE",
         anchor="w").pack(side=tk.BOTTOM, fill=tk.X)

# Main layout
main = tk.Frame(root, bg="#0f1720")
main.pack(fill=tk.BOTH, expand=True)

sidebar = tk.Frame(main, bg="#081025", width=250)
sidebar.pack(side=tk.LEFT, fill=tk.Y)

content = tk.Frame(main, bg="#071029")
content.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)

# Pages
home = tk.Frame(content, bg="#071029")
students = tk.Frame(content, bg="#071029")
attendance = tk.Frame(content, bg="#071029")
settings = tk.Frame(content, bg="#071029")

for p in (home, students, attendance, settings):
    p.place(relx=0, rely=0, relwidth=1, relheight=1)

def show(page):
    page.tkraise()
    if page is students:
        update_student_list()
    if page is attendance:
        load_marked_today()
        show_attendance()

# Sidebar buttons
def nav(text, cmd):
    tk.Button(sidebar, text=text, command=cmd,
              bg="#081025", fg="#DDE9FF",
              bd=0, anchor="w", padx=20, pady=12).pack(fill=tk.X, pady=2)

nav("🏠 Home", lambda: show(home))
nav("👥 Students", lambda: show(students))
nav("🧾 Attendance", lambda: show(attendance))
nav("⚙️ Settings", lambda: show(settings))
nav("🚪 Exit", close_app)

# ---------------- HOME PAGE ----------------
video_label = tk.Label(home, bg="black", bd=3, relief=tk.GROOVE)
video_label.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

btns = tk.Frame(home, bg="#071029")
btns.pack()

ttk.Button(btns, text="Start Camera", command=start_camera).grid(row=0, column=0, padx=6)
ttk.Button(btns, text="Stop Camera", command=stop_camera).grid(row=0, column=1, padx=6)
ttk.Button(btns, text="Add Student", command=start_capture).grid(row=0, column=2, padx=6)

# ---------------- STUDENTS PAGE ----------------
tk.Label(students, text="Registered Students",
         bg="#071029", fg="#DDE9FF",
         font=("Segoe UI", 16, "bold")).pack(padx=12, pady=10, anchor="w")

student_tree = ttk.Treeview(students, columns=("Name"), show="headings", height=18)
student_tree.heading("Name", text="Student Name")
student_tree.pack(fill=tk.BOTH, expand=True, padx=12)

st_btns = tk.Frame(students, bg="#071029")
st_btns.pack(pady=6)

ttk.Button(st_btns, text="Refresh", command=update_student_list).pack(side=tk.LEFT, padx=6)
ttk.Button(st_btns, text="Delete", command=delete_student).pack(side=tk.LEFT, padx=6)

# ---------------- ATTENDANCE PAGE ----------------
tk.Label(attendance, text="Attendance Records",
         bg="#071029", fg="#DDE9FF",
         font=("Segoe UI", 16, "bold")).pack(padx=12, pady=10, anchor="w")

search = tk.Frame(attendance, bg="#071029")
search.pack(fill=tk.X, padx=12)

search_name_var = tk.StringVar()
search_date_var = tk.StringVar()

tk.Label(search, text="Name:", bg="#071029", fg="#C7D2FE").pack(side=tk.LEFT)
tk.Entry(search, textvariable=search_name_var).pack(side=tk.LEFT, padx=6)

tk.Label(search, text="Date:", bg="#071029", fg="#C7D2FE").pack(side=tk.LEFT, padx=8)
tk.Entry(search, textvariable=search_date_var, width=12).pack(side=tk.LEFT)

ttk.Button(search, text="Search", command=show_attendance).pack(side=tk.LEFT, padx=6)
ttk.Button(search, text="Clear", command=clear_filters).pack(side=tk.LEFT, padx=6)

attendance_tree = ttk.Treeview(attendance,
                               columns=("Name", "Date", "Time"),
                               show="headings", height=20)
for col in ("Name", "Date", "Time"):
    attendance_tree.heading(col, text=col)

attendance_tree.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)

# ---------------- SETTINGS PAGE ----------------
tk.Label(settings, text="Settings",
         bg="#071029", fg="#DDE9FF",
         font=("Segoe UI", 16, "bold")).pack(pady=10)

s_tools = tk.Frame(settings, bg="#071029")
s_tools.pack(pady=6)

ttk.Button(s_tools, text="Retrain Recognizer", command=train_recognizer_threadsafe).pack(pady=5)
ttk.Button(s_tools, text="Load Recognizer", command=load_trained_recognizer).pack(pady=5)
ttk.Button(s_tools, text="Refresh Attendance", command=show_attendance).pack(pady=5)
ttk.Button(s_tools, text="Refresh Students", command=update_student_list).pack(pady=5)

# FINAL INIT
show(home)
root.protocol("WM_DELETE_WINDOW", close_app)
root.mainloop()

